import pandas as pd
from openpyxl import load_workbook

class FinancialModelEnumerator:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        self.sheets = self.workbook.sheetnames
    
    def enumerate_cells(self, sheet_name):
        sheet = self.workbook[sheet_name]
        for row in sheet.iter_rows(values_only=True):
            print(row)  # Print all cells in each row
    
    def identify_key_metrics(self, sheet_name, keywords):
        sheet = self.workbook[sheet_name]
        key_metrics = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in keywords:
                    key_metrics[cell.value] = (cell.row, cell.column)
        return key_metrics
    
    def update_values(self, sheet_name, key_metric, new_value):
        sheet = self.workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == key_metric:
                    target_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    target_cell.value = new_value
                    print(f"Updated {key_metric} to {new_value} in row {cell.row}")
                    break

    def save(self, new_file_path=None):
        if new_file_path:
            self.workbook.save(new_file_path)
        else:
            self.workbook.save(self.file_path)

# Usage
enumerator = FinancialModelEnumerator('financial_model.xlsx')
enumerator.enumerate_cells('Sheet1')
key_metrics = enumerator.identify_key_metrics('Sheet1', ['Revenue', 'Expenses'])
enumerator.update_values('Sheet1', 'Revenue', 100000)
enumerator.save('updated_financial_model.xlsx')
